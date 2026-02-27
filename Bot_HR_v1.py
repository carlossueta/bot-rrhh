"""
Bot de WhatsApp para Callbell
Requiere: pip install flask openpyxl requests gunicorn
"""

import re
import os
import json
import openpyxl
from flask import Flask, request, jsonify
from datetime import datetime
import requests

app = Flask(__name__)

# ─── CONFIGURACIÓN ───────────────────────────────────────────────────────────
CALLBELL_API_TOKEN = os.environ.get("CALLBELL_API_TOKEN", "bDfAzWJk9pMj5y8Y3U4RzuzzU9WSXYea.dd284001db7b17a762da3c3ce9edee1e8b98117b69b5b97dd8c484a7e98499cc")
EXCEL_FILE = r"c:\Users\carlo\OneDrive\Desktop\Python\BOT HR\empleados.xlsx"

# ─── SESIONES EN MEMORIA ─────────────────────────────────────────────────────
sessions = {}

# ─── CARGA DEL EXCEL ─────────────────────────────────────────────────────────
def cargar_empleados():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    empleados = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        telefono = re.sub(r'\D', '', str(row[0]).strip())
        dni      = str(row[1]).strip()
        nombre   = str(row[2]).strip()
        empleados[(telefono, dni)] = nombre
    return empleados

# ─── VALIDACIONES ────────────────────────────────────────────────────────────
def validar_formato_dni(dni):
    return bool(re.fullmatch(r'\d{1,8}', dni.strip()))

def validar_identidad(telefono, dni):
    empleados = cargar_empleados()
    telefono_limpio = re.sub(r'\D', '', telefono)
    return empleados.get((telefono_limpio, dni.strip()))

# ─── ENVÍO DE MENSAJES ───────────────────────────────────────────────────────
def enviar_mensaje(telefono, texto):
    url = "https://api.callbell.eu/v1/messages/send"
    headers = {
        "Authorization": f"Bearer {CALLBELL_API_TOKEN}",
        "Content-Type": "application/json"
    }
    payload = {
        "to": telefono,
        "from": "whatsapp",
        "type": "text",
        "content": {"text": texto}
    }
    r = requests.post(url, headers=headers, json=payload)
    print(f"Envío mensaje → {r.status_code}: {r.text}")

def enviar_botones(telefono, texto, botones):
    url = "https://api.callbell.eu/v1/messages/send"
    headers = {
        "Authorization": f"Bearer {CALLBELL_API_TOKEN}",
        "Content-Type": "application/json"
    }
    if len(botones) <= 3:
        payload = {
            "to": telefono,
            "from": "whatsapp",
            "type": "interactive",
            "body": {
                "type": "button",
                "body": {"text": texto},
                "action": {
                    "buttons": [
                        {"type": "reply", "reply": {"id": str(i+1), "title": b}}
                        for i, b in enumerate(botones)
                    ]
                }
            }
        }
    else:
        rows = [{"id": str(i+1), "title": b} for i, b in enumerate(botones)]
        payload = {
            "to": telefono,
            "from": "whatsapp",
            "type": "interactive",
            "body": {
                "type": "list",
                "body": {"text": texto},
                "action": {
                    "button": "Ver opciones",
                    "sections": [{"title": "Menú", "rows": rows}]
                }
            }
        }
    r = requests.post(url, headers=headers, json=payload)
    print(f"Envío botones → {r.status_code}: {r.text}")

# ─── MENSAJES POR TIPO DE AUSENCIA ───────────────────────────────────────────
MENSAJES_AUSENCIA = {
    "1": (
        "Por favor adjuntá el certificado o documentación por este medio.\n\n"
        "⚠️ *Importante:*\n"
        "• La recepción del certificado no implica la justificación automática.\n"
        "• Toda documentación es evaluada por Auditoría de RH.\n"
        "• El original debe presentarse dentro de las *48 horas* una vez reincorporado/a.\n"
        "• En licencias prolongadas (+72 hs), Auditoría puede solicitar verificación en centro médico.\n"
        "• RH podrá enviar médico a domicilio al último domicilio declarado."
    ),
    "2": (
        "Por favor adjuntá el certificado o documentación por este medio.\n\n"
        "⚠️ *Importante:*\n"
        "• La recepción del certificado no implica la justificación automática.\n"
        "• Toda documentación es evaluada por Auditoría de RH.\n"
        "• El original debe presentarse dentro de las *48 horas* una vez reincorporado/a.\n"
        "• Las ausencias por enfermedad familiar son justificadas, pero no pagas.\n"
        "• Ante cualquier duda, comunicate con RH."
    ),
    "3": (
        "Podés adjuntar la documentación correspondiente por este medio.\n\n"
        "⚠️ *Importante:*\n"
        "• La recepción del documento no implica la justificación automática.\n"
        "• Auditoría de RH lo revisará.\n"
        "• El original debe presentarse dentro de las *48 horas* una vez reincorporado/a.\n\n"
        "_Si no tenés documentación, respondé *No tengo*._"
    ),
    "4": (
        "Acompañamos este momento y lamentamos profundamente tu pérdida. 🕊️\n\n"
        "Por favor, cuando te sea posible, enviá la documentación correspondiente "
        "para registrar la ausencia.\n\n"
        "Si necesitás hablar con alguien de RH, estamos a disposición."
    ),
    "5": (
        "Si sufriste un accidente yendo o volviendo del trabajo, debés avisar *EN EL MOMENTO*.\n\n"
        "1️⃣ Avisá inmediatamente a tu supervisor.\n"
        "2️⃣ Llamá a Federación Patronal (CAP):\n"
        "📞 *0810-333-0096*\n\n"
        "Tené a mano:\n"
        "• DNI\n"
        "• Domicilio y teléfono\n"
        "• Fecha y hora del accidente\n"
        "• Lugar y cómo ocurrió\n"
        "• Si intervino policía o SAME\n"
        "• Nombre de quien realiza la denuncia\n\n"
        "La ART te indicará el prestador médico o traslado.\n"
        "El empleador debe entregarte el *Formulario de Denuncia Administrativa*.\n\n"
        "Por favor adjuntá la documentación de respaldo por este medio cuando puedas."
    ),
}

# ─── MENÚS ───────────────────────────────────────────────────────────────────
MENU_PRINCIPAL = [
    "1) Avisar ausencias",
    "2) Consultas",
    "3) Beneficios",
    "4) Actualizar datos",
    "5) Otros temas"
]

MENU_AUSENCIAS = [
    "1) Enfermedad propia",
    "2) Enfermedad familiar",
    "3) Problema personal",
    "4) Fallecimiento",
    "5) Accidente in itinere",
    "6) Volver al menú anterior"
]

def mostrar_menu_principal(conv_id, nombre):
    enviar_botones(conv_id, f"¿En qué puedo ayudarte, {nombre}?", MENU_PRINCIPAL)

def mostrar_menu_ausencias(conv_id, nombre):
    enviar_botones(conv_id, f"{nombre}, seleccioná el motivo de la ausencia:", MENU_AUSENCIAS)

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def extraer_opcion(msg):
    msg = msg.strip()
    match = re.match(r'^(\d+)', msg)
    return match.group(1) if match else msg.lower()

def registrar_ausencia(telefono, dni, nombre, motivo):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if "Ausencias" not in wb.sheetnames:
        ws = wb.create_sheet("Ausencias")
        ws.append(["Fecha", "Teléfono", "DNI", "Nombre", "Motivo"])
    else:
        ws = wb["Ausencias"]
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M"), telefono, dni, nombre, motivo])
    wb.save(EXCEL_FILE)

def registrar_archivo(telefono, dni, nombre, motivo, url_archivo):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if "Documentos" not in wb.sheetnames:
        ws = wb.create_sheet("Documentos")
        ws.append(["Fecha", "Teléfono", "DNI", "Nombre", "Motivo", "URL Archivo"])
    else:
        ws = wb["Documentos"]
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M"), telefono, dni, nombre, motivo, url_archivo])
    wb.save(EXCEL_FILE)

# ─── MÁQUINA DE ESTADOS ──────────────────────────────────────────────────────
def procesar_mensaje(telefono, mensaje, conv_id):
    msg    = mensaje.strip()
    sesion = sessions.get(telefono, {})
    estado = sesion.get("estado", "inicio")

    # ── INICIO ──────────────────────────────────────────────────────────────
    if estado == "inicio":
        sessions[telefono] = {"estado": "esperando_dni"}
        enviar_mensaje(conv_id,
            "¡Hola! Soy Cristobal, tu asistente virtual de RRHH 👋\n"
            "Por favor, ingresá tu *DNI* (solo números, sin puntos):"
        )

    # ── ESPERANDO DNI ────────────────────────────────────────────────────────
    elif estado == "esperando_dni":
        if not validar_formato_dni(msg):
            enviar_mensaje(conv_id,
                "❌ El formato del DNI no es correcto.\n"
                "Debe contener solo números, sin puntos (máximo 8 dígitos).\n"
                "Por favor, ingresalo nuevamente:"
            )
            return

        nombre = validar_identidad(telefono, msg)
        if not nombre:
            enviar_mensaje(conv_id,
                "⚠️ No pudimos validar tus datos.\n"
                "Verificá que el DNI sea correcto o comunicate con RRHH."
            )
            sessions[telefono]["estado"] = "esperando_dni"
            return

        sessions[telefono].update({
            "estado": "menu_principal",
            "dni": msg,
            "nombre": nombre
        })
        mostrar_menu_principal(conv_id, nombre)

    # ── MENÚ PRINCIPAL ───────────────────────────────────────────────────────
    elif estado == "menu_principal":
        opcion = extraer_opcion(msg)
        nombre = sessions[telefono]["nombre"]

        if opcion == "1":
            sessions[telefono]["estado"] = "menu_ausencias"
            mostrar_menu_ausencias(conv_id, nombre)
        elif opcion in ("2", "3", "4", "5"):
            # Secciones a implementar próximamente
            labels = {"2": "Consultas", "3": "Beneficios", "4": "Actualizar datos", "5": "Otros temas"}
            enviar_mensaje(conv_id,
                f"{nombre}, la sección *{labels[opcion]}* estará disponible próximamente.\n"
                "Enviá *menú* para volver al inicio."
            )
        else:
            mostrar_menu_principal(conv_id, nombre)

    # ── MENÚ AUSENCIAS ───────────────────────────────────────────────────────
    elif estado == "menu_ausencias":
        opcion = extraer_opcion(msg)
        nombre = sessions[telefono]["nombre"]

        MOTIVOS = {
            "1": "Enfermedad propia",
            "2": "Enfermedad familiar",
            "3": "Problema personal",
            "4": "Fallecimiento",
            "5": "Accidente in itinere"
        }

        if opcion == "6":
            sessions[telefono]["estado"] = "menu_principal"
            mostrar_menu_principal(conv_id, nombre)
        elif opcion in MOTIVOS:
            motivo = MOTIVOS[opcion]
            registrar_ausencia(telefono, sessions[telefono]["dni"], nombre, motivo)
            sessions[telefono]["estado"]  = f"esperando_archivo_{opcion}"
            sessions[telefono]["motivo"]  = motivo
            enviar_mensaje(conv_id, MENSAJES_AUSENCIA[opcion])
        else:
            mostrar_menu_ausencias(conv_id, nombre)

    # ── ESPERANDO ARCHIVO ────────────────────────────────────────────────────
    elif estado.startswith("esperando_archivo_"):
        nombre       = sessions[telefono]["nombre"]
        motivo       = sessions[telefono]["motivo"]
        tipo_msg     = sessions[telefono].get("tipo_mensaje", "text")
        opcion_actual = estado.replace("esperando_archivo_", "")

        if tipo_msg in ("document", "image", "video"):
            url_archivo = sessions[telefono].get("url_archivo", "")
            registrar_archivo(telefono, sessions[telefono]["dni"], nombre, motivo, url_archivo)
            enviar_mensaje(conv_id,
                f"✅ Gracias, {nombre}. Tu documentación fue recibida correctamente "
                "y será evaluada por Auditoría de RH."
            )
            sessions[telefono]["estado"] = "menu_principal"
            mostrar_menu_principal(conv_id, nombre)

        elif opcion_actual == "3" and msg.lower() in ("no", "no tengo", "sin documentación", "omitir"):
            enviar_mensaje(conv_id,
                f"Entendido, {nombre}. Tu ausencia fue registrada.\n"
                "Recordá presentar el original dentro de las 48 hs de reincorporarte."
            )
            sessions[telefono]["estado"] = "menu_principal"
            mostrar_menu_principal(conv_id, nombre)

        else:
            enviar_mensaje(conv_id,
                "⏳ Estoy esperando que adjuntes el archivo (imagen, PDF o documento).\n"
                + ("Si no tenés documentación ahora, respondé *No tengo*." if opcion_actual == "3" else "")
            )

    # ── FALLBACK ─────────────────────────────────────────────────────────────
    else:
        nombre = sessions.get(telefono, {}).get("nombre", "")
        if msg.lower() in ("menú", "menu", "inicio"):
            sessions[telefono]["estado"] = "menu_principal"
            mostrar_menu_principal(conv_id, nombre)
        else:
            enviar_mensaje(conv_id,
                "No entendí tu mensaje. Enviá *menú* para volver al inicio."
            )

# ─── WEBHOOK ─────────────────────────────────────────────────────────────────
@app.route("/webhook", methods=["POST"])
def webhook():
    data = request.json
    print("Webhook recibido:", json.dumps(data, indent=2, ensure_ascii=False))

    try:
        payload      = data["payload"]

        # Ignorar mensajes enviados por el bot (status != "received")
        if payload.get("status") != "received":
            print("Mensaje propio ignorado.")
            return jsonify({"status": "ok"}), 200

        mensaje      = payload.get("text", "").strip()
        telefono     = re.sub(r'\D', '', payload["from"])
        contact_uuid = payload["contact"]["uuid"]  # guardamos por si lo necesitamos
        tipo_mensaje = payload.get("type", "text")

        if telefono in sessions:
            sessions[telefono]["tipo_mensaje"] = tipo_mensaje
            sessions[telefono]["url_archivo"]  = payload.get("mediaUrl", "")

        procesar_mensaje(telefono, mensaje, telefono)  # pasamos teléfono como identificador de canal

    except (KeyError, TypeError) as e:
        print(f"Error procesando webhook: {e}")

    return jsonify({"status": "ok"}), 200

# ─── HEALTH CHECK (para que Render sepa que el servicio está vivo) ────────────
@app.route("/", methods=["GET"])
def health():
    return jsonify({"status": "Bot RRHH activo ✅"}), 200

if __name__ == "__main__":
    app.run(port=5000, debug=True)