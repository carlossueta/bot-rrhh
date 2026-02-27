"""
Bot de WhatsApp con Twilio
Requiere: pip install flask openpyxl twilio
"""

import re
import os
import json
import openpyxl
from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse
from twilio.rest import Client
from datetime import datetime

app = Flask(__name__)

# ─── CONFIGURACIÓN ───────────────────────────────────────────────────────────
TWILIO_ACCOUNT_SID  = os.environ.get("TWILIO_ACCOUNT_SID", "ACf732147f9454c4a3aa5ed074bed9c5e8")
TWILIO_AUTH_TOKEN   = os.environ.get("TWILIO_AUTH_TOKEN", "f2b2b364807cd2da4c94713296765e26")
TWILIO_WA_NUMBER    = "whatsapp:+14155238886"  # Sandbox de Twilio (o tu número aprobado)
EXCEL_FILE = r"c:\Users\carlo\OneDrive\Desktop\Python\BOT HR\empleados.xlsx"

client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

# ─── SESIONES EN MEMORIA ─────────────────────────────────────────────────────
sessions = {}

# ─── EXCEL ───────────────────────────────────────────────────────────────────
def cargar_empleados():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    empleados = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        telefono = re.sub(r'\D', '', str(row[0]).strip())
        dni      = str(row[1]).strip()
        nombre   = str(row[2]).strip()
        empleados[(telefono, dni)] = nombre
    return empleados

def validar_formato_dni(dni):
    return bool(re.fullmatch(r'\d{1,8}', dni.strip()))

def validar_identidad(telefono, dni):
    return cargar_empleados().get((telefono, dni.strip()))

def registrar_ausencia(telefono, dni, nombre, motivo):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if "Ausencias" not in wb.sheetnames:
        ws = wb.create_sheet("Ausencias")
        ws.append(["Fecha", "Teléfono", "DNI", "Nombre", "Motivo"])
    else:
        ws = wb["Ausencias"]
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M"), telefono, dni, nombre, motivo])
    wb.save(EXCEL_FILE)

def registrar_archivo(telefono, dni, nombre, motivo, url_archivo):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if "Documentos" not in wb.sheetnames:
        ws = wb.create_sheet("Documentos")
        ws.append(["Fecha", "Teléfono", "DNI", "Nombre", "Motivo", "URL Archivo"])
    else:
        ws = wb["Documentos"]
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M"), telefono, dni, nombre, motivo, url_archivo])
    wb.save(EXCEL_FILE)

# ─── ENVÍO DE MENSAJES ───────────────────────────────────────────────────────
def enviar_mensaje(telefono, texto):
    """Envía mensaje de texto simple via Twilio."""
    msg = client.messages.create(
        from_=TWILIO_WA_NUMBER,
        to=f"whatsapp:{telefono}",
        body=texto
    )
    print(f"Mensaje enviado → SID: {msg.sid}")

def enviar_menu(telefono, texto, opciones):
    """
    WhatsApp no soporta botones nativos en Twilio sandbox.
    En producción (número aprobado) se pueden usar Interactive Messages.
    Por ahora enviamos lista numerada clara.
    """
    cuerpo = texto + "\n\n"
    for op in opciones:
        cuerpo += f"{op}\n"
    cuerpo += "\n_Respondé con el número de la opción._"
    enviar_mensaje(telefono, cuerpo)

# ─── MENÚS ───────────────────────────────────────────────────────────────────
MENU_PRINCIPAL = [
    "1) Avisar ausencias",
    "2) Consultas",
    "3) Beneficios",
    "4) Actualizar datos",
    "5) Otros temas"
]

MENU_AUSENCIAS = [
    "1) Enfermedad propia",
    "2) Enfermedad familiar",
    "3) Problema personal",
    "4) Fallecimiento",
    "5) Accidente in itinere",
    "6) Volver al menú anterior"
]

MOTIVOS = {
    "1": "Enfermedad propia",
    "2": "Enfermedad familiar",
    "3": "Problema personal",
    "4": "Fallecimiento",
    "5": "Accidente in itinere"
}

MENSAJES_AUSENCIA = {
    "1": (
        "Por favor adjuntá el certificado o documentación por este medio.\n\n"
        "⚠️ *Importante:*\n"
        "• La recepción del certificado no implica la justificación automática.\n"
        "• Toda documentación es evaluada por Auditoría de RH.\n"
        "• El original debe presentarse dentro de las *48 horas* una vez reincorporado/a.\n"
        "• En licencias prolongadas (+72 hs), Auditoría puede solicitar verificación en centro médico.\n"
        "• RH podrá enviar médico a domicilio al último domicilio declarado."
    ),
    "2": (
        "Por favor adjuntá el certificado o documentación por este medio.\n\n"
        "⚠️ *Importante:*\n"
        "• La recepción del certificado no implica la justificación automática.\n"
        "• Toda documentación es evaluada por Auditoría de RH.\n"
        "• El original debe presentarse dentro de las *48 horas* una vez reincorporado/a.\n"
        "• Las ausencias por enfermedad familiar son justificadas, pero no pagas.\n"
        "• Ante cualquier duda, comunicate con RH."
    ),
    "3": (
        "Podés adjuntar la documentación correspondiente por este medio.\n\n"
        "⚠️ *Importante:*\n"
        "• La recepción del documento no implica la justificación automática.\n"
        "• Auditoría de RH lo revisará.\n"
        "• El original debe presentarse dentro de las *48 horas* una vez reincorporado/a.\n\n"
        "_Si no tenés documentación, respondé *No tengo*._"
    ),
    "4": (
        "Acompañamos este momento y lamentamos profundamente tu pérdida. 🕊️\n\n"
        "Por favor, cuando te sea posible, enviá la documentación correspondiente "
        "para registrar la ausencia.\n\n"
        "Si necesitás hablar con alguien de RH, estamos a disposición."
    ),
    "5": (
        "Si sufriste un accidente yendo o volviendo del trabajo, debés avisar *EN EL MOMENTO*.\n\n"
        "1️⃣ Avisá inmediatamente a tu supervisor.\n"
        "2️⃣ Llamá a Federación Patronal (CAP):\n"
        "📞 *0810-333-0096*\n\n"
        "Tené a mano:\n"
        "• DNI\n"
        "• Domicilio y teléfono\n"
        "• Fecha y hora del accidente\n"
        "• Lugar y cómo ocurrió\n"
        "• Si intervino policía o SAME\n"
        "• Nombre de quien realiza la denuncia\n\n"
        "La ART te indicará el prestador médico o traslado.\n"
        "El empleador debe entregarte el *Formulario de Denuncia Administrativa*.\n\n"
        "Por favor adjuntá la documentación de respaldo por este medio cuando puedas."
    ),
}

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def extraer_opcion(msg):
    match = re.match(r'^(\d+)', msg.strip())
    return match.group(1) if match else msg.strip().lower()

# ─── MÁQUINA DE ESTADOS ──────────────────────────────────────────────────────
def procesar_mensaje(telefono, mensaje, num_media, media_url):
    msg    = mensaje.strip()
    sesion = sessions.get(telefono, {})
    estado = sesion.get("estado", "inicio")

    # ── INICIO ──────────────────────────────────────────────────────────────
    if estado == "inicio":
        sessions[telefono] = {"estado": "esperando_dni"}
        enviar_mensaje(telefono,
            "¡Hola! Soy Cristobal, el asistente virtual de RRHH 👋\n"
            "Por favor, ingresá tu *DNI* (solo números, sin puntos):"
        )

    # ── ESPERANDO DNI ────────────────────────────────────────────────────────
    elif estado == "esperando_dni":
        if not validar_formato_dni(msg):
            enviar_mensaje(telefono,
                "❌ El formato del DNI no es correcto.\n"
                "Debe contener solo números, sin puntos (máximo 8 dígitos).\n"
                "Por favor, ingresalo nuevamente:"
            )
            return
        nombre = validar_identidad(telefono, msg)
        if not nombre:
            enviar_mensaje(telefono,
                "⚠️ No pudimos validar tus datos.\n"
                "Verificá que el DNI sea correcto o comunicate con RRHH."
            )
            sessions[telefono]["estado"] = "esperando_dni"
            return
        sessions[telefono].update({"estado": "menu_principal", "dni": msg, "nombre": nombre})
        enviar_menu(telefono, f"¡Hola, {nombre}! ¿En qué puedo ayudarte?", MENU_PRINCIPAL)

    # ── MENÚ PRINCIPAL ───────────────────────────────────────────────────────
    elif estado == "menu_principal":
        opcion = extraer_opcion(msg)
        nombre = sessions[telefono]["nombre"]
        if opcion == "1":
            sessions[telefono]["estado"] = "menu_ausencias"
            enviar_menu(telefono, f"{nombre}, seleccioná el motivo de la ausencia:", MENU_AUSENCIAS)
        elif opcion in ("2", "3", "4", "5"):
            labels = {"2": "Consultas", "3": "Beneficios", "4": "Actualizar datos", "5": "Otros temas"}
            enviar_mensaje(telefono, f"{nombre}, la sección *{labels[opcion]}* estará disponible próximamente.\nEnviá *menú* para volver al inicio.")
        else:
            enviar_menu(telefono, f"¿En qué puedo ayudarte, {nombre}?", MENU_PRINCIPAL)

    # ── MENÚ AUSENCIAS ───────────────────────────────────────────────────────
    elif estado == "menu_ausencias":
        opcion = extraer_opcion(msg)
        nombre = sessions[telefono]["nombre"]
        if opcion == "6":
            sessions[telefono]["estado"] = "menu_principal"
            enviar_menu(telefono, f"¿En qué puedo ayudarte, {nombre}?", MENU_PRINCIPAL)
        elif opcion in MOTIVOS:
            motivo = MOTIVOS[opcion]
            registrar_ausencia(telefono, sessions[telefono]["dni"], nombre, motivo)
            sessions[telefono]["estado"]  = f"esperando_archivo_{opcion}"
            sessions[telefono]["motivo"]  = motivo
            enviar_mensaje(telefono, MENSAJES_AUSENCIA[opcion])
        else:
            enviar_menu(telefono, f"{nombre}, seleccioná el motivo de la ausencia:", MENU_AUSENCIAS)

    # ── ESPERANDO ARCHIVO ────────────────────────────────────────────────────
    elif estado.startswith("esperando_archivo_"):
        nombre        = sessions[telefono]["nombre"]
        motivo        = sessions[telefono]["motivo"]
        opcion_actual = estado.replace("esperando_archivo_", "")

        if num_media and int(num_media) > 0:
            registrar_archivo(telefono, sessions[telefono]["dni"], nombre, motivo, media_url)
            enviar_mensaje(telefono,
                f"✅ Gracias, {nombre}. Tu documentación fue recibida y será evaluada por Auditoría de RH."
            )
            sessions[telefono]["estado"] = "menu_principal"
            enviar_menu(telefono, f"¿En qué puedo ayudarte, {nombre}?", MENU_PRINCIPAL)
        elif opcion_actual == "3" and msg.lower() in ("no", "no tengo", "sin documentación", "omitir"):
            enviar_mensaje(telefono,
                f"Entendido, {nombre}. Tu ausencia fue registrada.\n"
                "Recordá presentar el original dentro de las 48 hs de reincorporarte."
            )
            sessions[telefono]["estado"] = "menu_principal"
            enviar_menu(telefono, f"¿En qué puedo ayudarte, {nombre}?", MENU_PRINCIPAL)
        else:
            enviar_mensaje(telefono,
                "⏳ Estoy esperando que adjuntes el archivo (imagen o PDF).\n"
                + ("Si no tenés documentación, respondé *No tengo*." if opcion_actual == "3" else "")
            )

    # ── FALLBACK ─────────────────────────────────────────────────────────────
    else:
        nombre = sessions.get(telefono, {}).get("nombre", "")
        if msg.lower() in ("menú", "menu", "inicio"):
            sessions[telefono]["estado"] = "menu_principal"
            enviar_menu(telefono, f"¿En qué puedo ayudarte, {nombre}?", MENU_PRINCIPAL)
        else:
            enviar_mensaje(telefono, "No entendí tu mensaje. Enviá *menú* para volver al inicio.")

# ─── WEBHOOK TWILIO ───────────────────────────────────────────────────────────
@app.route("/webhook", methods=["POST"])
def webhook():
    telefono  = re.sub(r'\D', '', request.form.get("From", ""))
    mensaje   = request.form.get("Body", "").strip()
    num_media = request.form.get("NumMedia", "0")
    media_url = request.form.get("MediaUrl0", "")

    print(f"Mensaje de {telefono}: '{mensaje}' | Media: {num_media}")
    procesar_mensaje(telefono, mensaje, num_media, media_url)

    # Twilio espera una respuesta TwiML (puede estar vacía)
    return str(MessagingResponse()), 200

# ─── HEALTH CHECK ────────────────────────────────────────────────────────────
@app.route("/", methods=["GET"])
def health():
    return "Bot RRHH con Twilio ✅", 200

if __name__ == "__main__":
    app.run(port=5000, debug=True)